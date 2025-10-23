"""
Result Exporter - Export exam results to PDF and Excel formats
"""

import csv
import json
from io import BytesIO, StringIO
from datetime import datetime
from typing import List, Dict, Any, Tuple
import logging

logger = logging.getLogger(__name__)

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


class ResultExporter:
    """Export exam results to various formats"""
    
    @staticmethod
    def export_to_csv(
        results: List[Dict[str, Any]],
        filename: str = None
    ) -> Tuple[BytesIO, str]:
        """
        Export results to CSV format
        
        Args:
            results: List of result dictionaries
            filename: Optional filename
        
        Returns:
            (BytesIO object, filename)
        """
        if not filename:
            filename = f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        output = StringIO()
        
        if not results:
            return BytesIO(b''), filename
        
        # Get all unique keys
        fieldnames = set()
        for result in results:
            fieldnames.update(result.keys())
        fieldnames = sorted(list(fieldnames))
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)
        
        # Convert to BytesIO
        csv_bytes = BytesIO(output.getvalue().encode('utf-8'))
        
        logger.info(f"Exported {len(results)} results to CSV")
        return csv_bytes, filename
    
    @staticmethod
    def export_to_excel(
        results: List[Dict[str, Any]],
        filename: str = None
    ) -> Tuple[BytesIO, str]:
        """
        Export results to Excel format
        
        Args:
            results: List of result dictionaries
            filename: Optional filename
        
        Returns:
            (BytesIO object, filename)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return ResultExporter.export_to_csv(results, filename)
        
        if not filename:
            filename = f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        
        if not results:
            return BytesIO(b''), filename
        
        # Get headers
        headers = sorted(list(set().union(*(d.keys() for d in results))))
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_idx, result in enumerate(results, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = result.get(header, '')
                
                # Format score columns
                if 'score' in header.lower() or 'percentage' in header.lower():
                    if isinstance(value, (int, float)):
                        cell.value = value
                        cell.number_format = '0.00'
                else:
                    cell.value = value
                
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Exported {len(results)} results to Excel")
        return output, filename
    
    @staticmethod
    def export_to_pdf(
        results: List[Dict[str, Any]],
        title: str = "Exam Results",
        filename: str = None
    ) -> Tuple[BytesIO, str]:
        """
        Export results to PDF format
        
        Args:
            results: List of result dictionaries
            title: Report title
            filename: Optional filename
        
        Returns:
            (BytesIO object, filename)
        """
        if not HAS_REPORTLAB:
            logger.warning("reportlab not installed, falling back to CSV")
            return ResultExporter.export_to_csv(results, filename)
        
        if not filename:
            filename = f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Summary
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=20
        )
        summary_text = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {len(results)}"
        elements.append(Paragraph(summary_text, summary_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        if not results:
            elements.append(Paragraph("No results to display", styles['Normal']))
        else:
            # Create table
            headers = sorted(list(set().union(*(d.keys() for d in results))))
            
            # Limit columns for PDF readability
            if len(headers) > 8:
                headers = headers[:8]
            
            table_data = [headers]
            for result in results:
                row = [str(result.get(h, ''))[:50] for h in headers]
                table_data.append(row)
            
            table = Table(table_data, colWidths=[1.2 * inch] * len(headers))
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        logger.info(f"Exported {len(results)} results to PDF")
        return output, filename
    
    @staticmethod
    def export_to_json(
        results: List[Dict[str, Any]],
        filename: str = None
    ) -> Tuple[BytesIO, str]:
        """
        Export results to JSON format
        
        Args:
            results: List of result dictionaries
            filename: Optional filename
        
        Returns:
            (BytesIO object, filename)
        """
        if not filename:
            filename = f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        data = {
            'exported_at': datetime.now().isoformat(),
            'total_records': len(results),
            'results': results
        }
        
        json_str = json.dumps(data, indent=2, default=str)
        json_bytes = BytesIO(json_str.encode('utf-8'))
        
        logger.info(f"Exported {len(results)} results to JSON")
        return json_bytes, filename
    
    @staticmethod
    def export_summary_report(
        results: List[Dict[str, Any]],
        format: str = 'pdf'
    ) -> Tuple[BytesIO, str]:
        """
        Export summary report with statistics
        
        Args:
            results: List of result dictionaries
            format: Export format (pdf, csv, excel, json)
        
        Returns:
            (BytesIO object, filename)
        """
        # Calculate statistics
        if not results:
            summary = {'message': 'No results to summarize'}
        else:
            scores = [r.get('score', 0) for r in results if isinstance(r.get('score'), (int, float))]
            summary = {
                'total_submissions': len(results),
                'average_score': sum(scores) / len(scores) if scores else 0,
                'highest_score': max(scores) if scores else 0,
                'lowest_score': min(scores) if scores else 0,
                'pass_count': sum(1 for s in scores if s >= 70),
                'pass_rate': (sum(1 for s in scores if s >= 70) / len(scores) * 100) if scores else 0
            }
        
        summary_results = [summary]
        
        if format == 'pdf':
            return ResultExporter.export_to_pdf(summary_results, "Exam Results Summary")
        elif format == 'excel':
            return ResultExporter.export_to_excel(summary_results)
        elif format == 'json':
            return ResultExporter.export_to_json(summary_results)
        else:
            return ResultExporter.export_to_csv(summary_results)

